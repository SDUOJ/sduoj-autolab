"""Automation task for generating summary reports and exporting to Excel."""

from __future__ import annotations

import asyncio
import io
import json
import logging
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field

from model.answer_sheet import answerSheetModel
from model.problem_set import problemSetModel
from db import ProblemGroup, ProblemSet
from sduojApi import getGroupMember, getGroupName, uploadFiles

from .base import BaseAutoTask, register_task

logger = logging.getLogger(__name__)


# --- Payload Schema ---

class SummaryReportPayload(BaseModel):
    """Payload for summary report generation task."""
    groupId: int
    psids: Optional[List[int]] = Field(default=None)
    userId: int  # User ID for file upload attribution


# --- Helper Functions ---

def number_to_excel_column(n: int) -> str:
    """Converts a 1-based integer to an Excel column name (e.g., 1 -> A, 27 -> AA)."""
    column_name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        column_name = chr(65 + remainder) + column_name
    return column_name


def gc(row: int, col: int) -> str:
    """Returns the Excel cell coordinate string (e.g., (1, 1) -> "A1")."""
    return f"{number_to_excel_column(col)}{row}"


def sc(ws, start: tuple, value, font: Font):
    """Sets a value with style in a single cell."""
    alignment = Alignment(horizontal='center', vertical='center')
    cell = ws[gc(*start)]
    cell.value = value
    cell.alignment = alignment
    cell.font = font


def scm(ws, start: tuple, end: tuple, value, font: Font):
    """Sets a value with style in merged cells."""
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell = ws[gc(*start)]
    cell.value = value
    cell.alignment = alignment
    cell.font = font
    if start != end:
        ws.merge_cells(f"{gc(*start)}:{gc(*end)}")


def safe_int(value, default=0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def build_submission_from_summary(score: Any) -> Dict[str, Any]:
    """Convert summary data into sheet-friendly fields."""
    return {
        'score': score or 0
    }


@register_task("summary_report")
class SummaryReportTask(BaseAutoTask):
    """Generate comprehensive summary report and export to Excel."""

    def run(self) -> Dict[str, Any]:
        payload = self._parse_payload()
        logger.info(
            "Summary report task started for group %s with psids: %s",
            payload.groupId,
            payload.psids,
        )
        
        service = _SummaryReportService(payload, self.raw_task.get("task_id"))
        result = service.execute()
        
        logger.info("Summary report task completed for group %s", payload.groupId)
        return result

    def _parse_payload(self) -> SummaryReportPayload:
        if not isinstance(self.payload, dict):
            raise ValueError("Payload must be a dict")
        return SummaryReportPayload.parse_obj(self.payload)


class _SummaryReportService:
    """Service class handling the actual report generation logic."""

    def __init__(self, payload: SummaryReportPayload, task_id: Optional[str]):
        self.payload = payload
        self.task_id = task_id
        
        self.answer_model = answerSheetModel()
        self.ps_model = problemSetModel()

    def execute(self) -> Dict[str, Any]:
        """Main execution flow."""
        created_loop = False
        try:
            try:
                loop = asyncio.get_event_loop()
            except RuntimeError:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                created_loop = True
            if loop.is_closed():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                created_loop = True
            try:
                return loop.run_until_complete(self._execute_inner())
            finally:
                if created_loop:
                    loop.close()
        finally:
            self._cleanup()

    def _cleanup(self):
        """Clean up resources."""
        for model in [self.answer_model, self.ps_model]:
            try:
                model.session.close()
            except Exception:
                pass

    def _get_problem_sets_by_group(
        self,
        group_id: int,
        psids: Optional[List[int]] = None
    ) -> List[ProblemSet]:
        """Fetch problem sets for a group, optionally filtered by psids."""
        query = self.ps_model.session.query(ProblemSet).filter(
            ProblemSet.groupId == group_id
        )
        if psids:
            query = query.filter(ProblemSet.psid.in_(psids))
        return query.all()

    def _get_problem_groups_batch(self, gids: List[int]) -> Dict[int, ProblemGroup]:
        """Batch fetch problem groups by IDs."""
        if not gids:
            return {}
        groups = self.ps_model.session.query(ProblemGroup).filter(
            ProblemGroup.gid.in_(gids)
        ).all()
        return {g.gid: g for g in groups}

    def _build_tag_structure(self, ps_objects: List[ProblemSet]) -> Dict[str, List[Dict[str, Any]]]:
        """Build hierarchical tag structure with weights from problem sets."""
        tag_structure: Dict[str, List[Dict[str, Any]]] = {}

        all_gids = set()
        ps_group_map = {}

        for ps_obj in ps_objects:
            try:
                group_info = json.loads(ps_obj.groupInfo) if isinstance(ps_obj.groupInfo, str) else ps_obj.groupInfo
            except Exception:
                group_info = []

            ps_group_map[ps_obj.psid] = group_info
            for g_info in group_info or []:
                if isinstance(g_info, dict) and 'gid' in g_info:
                    all_gids.add(g_info['gid'])

        group_objects = self._get_problem_groups_batch(list(all_gids))

        for ps_obj in ps_objects:
            ps_tag = ps_obj.tag or "未分类"
            ps_name = ps_obj.name
            ps_global_score = ps_obj.global_score or 1.0
            group_info = ps_group_map.get(ps_obj.psid, [])

            groups_detail = []
            for g_info in group_info or []:
                gid = g_info.get('gid')
                if not gid or gid not in group_objects:
                    continue

                g_obj = group_objects[gid]
                g_score = g_info.get('score', 1.0) or 1.0

                try:
                    problem_info = json.loads(g_obj.problemInfo) if isinstance(g_obj.problemInfo, str) else g_obj.problemInfo
                except Exception:
                    problem_info = []

                problems = []
                for p_info in problem_info or []:
                    problems.append({
                        'pid': p_info.get('pid'),
                        'score': p_info.get('score', 1.0) or 1.0,
                        'submitLimit': p_info.get('submitLimit', 0),
                        'antiCheatingRate': p_info.get('antiCheatingRate', 0.85),
                        'title': p_info.get('title'),
                        'name': p_info.get('name')
                    })

                groups_detail.append({
                    'gid': gid,
                    'name': g_obj.name,
                    'type': g_obj.type,
                    'score': g_score,
                    'problems': problems
                })

            tag_structure.setdefault(ps_tag, []).append({
                'psid': ps_obj.psid,
                'name': ps_name,
                'weight': ps_global_score,
                'groups': groups_detail
            })

        return tag_structure

    async def _execute_inner(self) -> Dict[str, Any]:
        """Internal execution logic."""
        # 1. Fetch group info and members
        group_title = await getGroupName(self.payload.groupId)
        group_info = await getGroupMember(self.payload.groupId)
        members = group_info.get("members", [])
        
        # Determine personnel to exclude
        exclude_usernames = {"superadmin"}
        
        # Add group owner/creator
        owner = group_info.get("username")
        if owner:
            exclude_usernames.add(owner)
        
        # Add current group's management group members
        current_manage_group_id = group_info.get("manageGroupId")
        if current_manage_group_id:
            current_manage_info = await getGroupMember(current_manage_group_id)
            for m_mem in current_manage_info.get("members", []):
                exclude_usernames.add(m_mem["username"])
        
        # Get all problem sets to collect their management groups
        ps_objects = self._get_problem_sets_by_group(
            self.payload.groupId,
            self.payload.psids
        )
        
        # Collect all unique management group IDs from problem sets
        ps_manage_group_ids = set()
        for ps_obj in ps_objects:
            if ps_obj.manageGroupId:
                ps_manage_group_ids.add(ps_obj.manageGroupId)
        
        # Fetch all management groups and add their members to exclusion list
        for mg_id in ps_manage_group_ids:
            try:
                mg_info = await getGroupMember(mg_id)
                for m_mem in mg_info.get("members", []):
                    exclude_usernames.add(m_mem["username"])
            except Exception as e:
                logger.warning(f"Failed to fetch management group {mg_id}: {e}")
        
        # Also exclude any member with admin-like roles in the current group
        admin_roles = {"admin", "manager", "teacher", "ta"}
        for mem in members:
            mem_roles = set(mem.get("roles", []))
            if mem_roles.intersection(admin_roles):
                exclude_usernames.add(mem["username"])

        # Filter out TAs, managers, creator, and superadmin
        member_list = [
            [mem["username"], mem.get("nickname", mem["username"])]
            for mem in members
            if mem["username"] not in exclude_usernames
        ]
        member_list.sort(key=lambda x: x[0]) # Sort by student ID
        
        if not member_list:
            raise ValueError("No valid members found in group")
        
        logger.info(f"Processing report for {len(member_list)} members")
        
        # 2. Fetch problem set data organized by tags (using ps_objects we already fetched)
        summary_data, tag_structure, ps_info_dict = await self._fetch_summary_data(ps_objects)
        
        # 3. Generate Excel with weights sheet and time-based scoring (async)
        excel_bytes = await self._generate_excel(
            group_title=group_title,
            member_list=member_list,
            summary_data=summary_data,
            tag_structure=tag_structure,
            ps_info_dict=ps_info_dict
        )
        
        # 4. Upload to file system
        file_info_list = await uploadFiles(
            files=[{
                "filename": f"{group_title}_成绩报告.xlsx",
                "content": excel_bytes,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }],
            userId=self.payload.userId
        )
        
        if not file_info_list:
            raise RuntimeError("Failed to upload Excel file")
        
        file_info = file_info_list[0]
        
        return {
            "fileId": file_info["id"],
            "fileName": file_info["name"],
            "size": file_info.get("size", 0),
            "message": f"成绩报告已生成，共 {len(member_list)} 名学生"
        }

    async def _fetch_summary_data(
        self,
        ps_objects: List
    ) -> tuple[Dict[int, Dict[str, Any]], Dict[str, List[Dict]], Dict[int, Dict[str, Any]]]:
        """
        Fetch summary data for all psids, organized by tags.
        Returns (summary_data, tag_structure, ps_info_dict)
        
        Args:
            ps_objects: List of ProblemSet objects already fetched
        """
        # Build tag structure with all weight information
        tag_structure = self._build_tag_structure(ps_objects)
        ps_info_dict: Dict[int, Dict[str, Any]] = {}

        for ps_obj in ps_objects:
            try:
                config = json.loads(ps_obj.config) if isinstance(ps_obj.config, str) else ps_obj.config
            except Exception:
                config = {}
            try:
                group_info = json.loads(ps_obj.groupInfo) if isinstance(ps_obj.groupInfo, str) else ps_obj.groupInfo
            except Exception:
                group_info = []
            ps_info_dict[ps_obj.psid] = {
                "tm_start": ps_obj.tm_start,
                "tm_end": ps_obj.tm_end,
                "config": config or {},
                "groupInfo": group_info or []
            }
        
        # Fetch detailed score data for each problem set
        summary_data = {}
        for ps_obj in ps_objects:
            psid = ps_obj.psid
            # Still use get_all_progress_cache for compatibility with existing data structure
            data = await self.answer_model.get_all_progress_cache(psid, 1)  # get_code=1 for full data
            summary_data[psid] = data
        
        return summary_data, tag_structure, ps_info_dict

    async def _generate_excel(
        self,
        group_title: str,
        member_list: List[List[str]],
        summary_data: Dict[int, Dict[str, Any]],
        tag_structure: Dict[str, List[Dict]],
        ps_info_dict: Dict[int, Dict[str, Any]]
    ) -> bytes:
        """Generate Excel file with summary scores."""
        wb = openpyxl.Workbook()
        
        # Create main summary sheet
        ws_summary = wb.active
        ws_summary.title = "总成绩"
        
        # Create weights sheet
        ws_weights = wb.create_sheet(title="权重配置")
        
        font10 = Font(name='Arial', size=10)
        font12b = Font(name='Arial', size=12, bold=True)
        font14b = Font(name='Arial', size=14, bold=True)
        font20b = Font(name='Arial', size=20, bold=True)
        
        # 1. Generate weights sheet
        self._create_weights_sheet(ws_weights, tag_structure, font10, font12b, font14b)
        
        # 2. Initialize student score structure
        student_scores = {}
        for uname, nick in member_list:
            student_scores[uname] = {
                "nickname": nick,
                "tag_totals": {},
                "scores": {},
                "total": 0
            }
        
        # Process data by tags
        sorted_tags = sorted(tag_structure.keys())
        tag_ps_structures = {}
        
        for tag in sorted_tags:
            tag_ps_structures[tag] = []
            for ps_info in tag_structure[tag]:
                psid = ps_info['psid']
                ps_name = ps_info['name']
                ps_meta = ps_info_dict.get(psid, {})
                group_info = ps_meta.get("groupInfo", [])
                gid_to_index = {}
                if isinstance(group_info, list):
                    for idx, group in enumerate(group_info):
                        if isinstance(group, dict) and 'gid' in group:
                            gid_to_index[safe_int(group['gid'], group['gid'])] = idx
                
                # Extract problem structure with pid
                problems = []
                for group_order_idx, group in enumerate(ps_info.get('groups', [])):
                    gid = safe_int(group.get('gid', group_order_idx), group_order_idx)
                    g_idx = gid_to_index.get(gid, group_order_idx)
                    group_name = group.get('name') or f"分组{g_idx + 1}"
                    for p_idx, prob in enumerate(group.get('problems', []), start=1):
                        pid = safe_int(prob.get('pid', 0), 0)
                        prob_key = f"{gid}-{pid}"
                        prob_name = prob.get('title') or prob.get('name') or f"{group_name}-{p_idx}"
                        problems.append({
                            'key': prob_key,
                            'gid': gid,
                            'pid': pid,
                            'name': prob_name,
                            'type': group.get('type', 0),
                            'group_idx': g_idx,
                            'problem_idx': p_idx - 1
                        })
                
                tag_ps_structures[tag].append((psid, ps_name, problems))
                
                # Initialize score structure for all students
                for uname in student_scores:
                    if psid not in student_scores[uname]["scores"]:
                        student_scores[uname]["scores"][psid] = {
                            "total": 0,
                            "submissions": {}
                        }
        
        # 3. Build score maps using summary data
        for tag in sorted_tags:
            for psid, ps_name, problems in tag_ps_structures[tag]:
                summary_rows = summary_data.get(psid, {}).get("data", [])
                summary_map = {
                    row.get("username"): row for row in summary_rows if row.get("username")
                }
                
                # For each student-problem pair, pull score from summary data
                for uname in student_scores:
                    summary_row = summary_map.get(uname, {})
                    for prob in problems:
                        key = prob['key']
                        score_key = f"{safe_int(prob.get('group_idx', 0), 0) + 1}-{safe_int(prob.get('problem_idx', 0), 0) + 1}"
                        score_entry = summary_row.get(score_key, {}) if isinstance(summary_row, dict) else {}
                        score_val = score_entry.get('s', 0) if isinstance(score_entry, dict) else 0
                        student_scores[uname]["scores"][psid]["submissions"][key] = build_submission_from_summary(
                            score_val
                        )
        
        # 4. Create summary sheet
        ps_weight_map = {
            ps_info['psid']: ps_info.get('weight', 0)
            for ps_list in tag_structure.values()
            for ps_info in ps_list
        }
        self._create_summary_sheet(ws_summary, group_title, member_list, student_scores,
                                  tag_ps_structures, sorted_tags, ps_weight_map,
                                  font10, font12b, font14b, font20b)
        
        # 5. Create detail sheet for each problem set
        for tag in sorted_tags:
            for psid, ps_name, problems in tag_ps_structures[tag]:
                sheet_name = self._sanitize_sheet_name(ps_name, psid)
                ws_detail = wb.create_sheet(title=sheet_name)
                
                self._create_detail_sheet(
                    ws_detail, ps_name, psid, member_list, student_scores,
                    problems,
                    font10, font12b, font14b
                )
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _sanitize_sheet_name(self, name: str, psid: int) -> str:
        """Sanitize sheet name to meet Excel requirements (max 31 chars, no special chars)."""
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '')
        
        # Truncate and add ID if too long
        if len(name) > 25:
            name = name[:25]
        return f"{name}_{psid}"[:31]
    
    def _create_weights_sheet(self, ws, tag_structure, font10, font12b, font14b):
        """Create the weights configuration sheet with hierarchical structure and formula-based weights."""
        # Headers
        sc(ws, (1, 1), "标签", font14b)
        sc(ws, (1, 2), "题单名称", font14b)
        sc(ws, (1, 3), "题单ID", font14b)
        sc(ws, (1, 4), "题单权重", font14b)
        sc(ws, (1, 5), "题组名称", font14b)
        sc(ws, (1, 6), "题组系数", font14b)
        sc(ws, (1, 7), "题组权重", font14b)
        sc(ws, (1, 8), "题目ID", font14b)
        sc(ws, (1, 9), "题目系数", font14b)
        sc(ws, (1, 10), "题目权重", font14b)
        sc(ws, (1, 11), "备注", font14b)
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 30
        
        row = 2
        for tag, ps_list in sorted(tag_structure.items()):
            tag_start_row = row
            
            for ps_info in ps_list:
                ps_start_row = row
                groups = ps_info.get('groups', [])
                ps_weight = ps_info['weight']
                
                if not groups:
                    # No groups
                    sc(ws, (row, 1), tag, font10)
                    sc(ws, (row, 2), ps_info['name'], font10)
                    sc(ws, (row, 3), ps_info['psid'], font10)
                    sc(ws, (row, 4), ps_weight, font10)
                    sc(ws, (row, 11), "无题组", font10)
                    row += 1
                    continue

                group_row_counts = []
                for g_info in groups:
                    problems = g_info.get('problems', [])
                    group_row_counts.append(max(1, len(problems)))

                ps_end_row = ps_start_row + sum(group_row_counts) - 1

                # Calculate sum of group coefficients for formulas
                for g_info, group_row_count in zip(groups, group_row_counts):
                    g_start_row = row
                    g_end_row = g_start_row + group_row_count - 1
                    problems = g_info.get('problems', [])
                    g_coef = g_info['score']

                    ps_weight_cell = f"D{ps_start_row}"
                    group_coef_cell = f"F{g_start_row}"
                    group_coef_sum_range = f"$F${ps_start_row}:$F${ps_end_row}"
                    problem_coef_sum_range = f"$I${g_start_row}:$I${g_end_row}"
                    group_weight_formula = f"={ps_weight_cell}*{group_coef_cell}/SUM({group_coef_sum_range})"
                    
                    if not problems:
                        sc(ws, (row, 1), tag, font10)
                        sc(ws, (row, 2), ps_info['name'], font10)
                        sc(ws, (row, 3), ps_info['psid'], font10)
                        sc(ws, (row, 4), ps_weight, font10)
                        sc(ws, (row, 5), g_info['name'], font10)
                        sc(ws, (row, 6), g_coef, font10)
                        # Group weight formula
                        cell = ws[gc(row, 7)]
                        cell.value = group_weight_formula
                        cell.font = font10
                        sc(ws, (row, 11), "无题目", font10)
                        row += 1
                        continue
                    
                    for p_info in problems:
                        p_coef = p_info['score']
                        
                        sc(ws, (row, 1), tag, font10)
                        sc(ws, (row, 2), ps_info['name'], font10)
                        sc(ws, (row, 3), ps_info['psid'], font10)
                        sc(ws, (row, 4), ps_weight, font10)
                        sc(ws, (row, 5), g_info['name'], font10)
                        sc(ws, (row, 6), g_coef, font10)
                        
                        # Group weight formula: ps_weight * (g_coef / sum_of_group_coefs_in_ps)
                        cell = ws[gc(row, 7)]
                        cell.value = group_weight_formula
                        cell.font = font10
                        
                        sc(ws, (row, 8), p_info['pid'], font10)
                        sc(ws, (row, 9), p_coef, font10)
                        
                        # Problem weight formula: g_weight * (p_coef / sum_of_prob_coefs_in_group)
                        # Need to identify rows with same gid under same psid
                        cell = ws[gc(row, 10)]
                        cell.value = f"=G{row}*I{row}/SUM({problem_coef_sum_range})"
                        cell.font = font10
                        
                        # Build remark
                        remark_parts = []
                        if p_info.get('submitLimit'):
                            remark_parts.append(f"提交限制:{p_info['submitLimit']}次")
                        if p_info.get('antiCheatingRate') and p_info['antiCheatingRate'] < 1.0:
                            remark_parts.append(f"查重阈值:{p_info['antiCheatingRate']}")
                        sc(ws, (row, 11), "; ".join(remark_parts) if remark_parts else "", font10)
                        row += 1
                    
                    # Merge group cells
                    if row > g_start_row + 1:
                        ws.merge_cells(f"E{g_start_row}:E{row-1}")
                        ws.merge_cells(f"F{g_start_row}:F{row-1}")
                
                # Merge problem set cells
                if row > ps_start_row + 1:
                    ws.merge_cells(f"B{ps_start_row}:B{row-1}")
                    ws.merge_cells(f"C{ps_start_row}:C{row-1}")
                    ws.merge_cells(f"D{ps_start_row}:D{row-1}")
            
            # Merge tag cells
            if row > tag_start_row:
                ws.merge_cells(f"A{tag_start_row}:A{row-1}")
    
    def _create_summary_sheet(self, ws, group_title, member_list, student_scores,
                            tag_ps_structures, sorted_tags, ps_weight_map,
                            font10, font12b, font14b, font20b):
        """Create the summary sheet with only totals and subtotals using formulas."""
        
        # Calculate total columns needed: Name, ID, Total, Tag totals + PS subtotals
        total_cols = 3  # Name, ID, Total
        for tag in sorted_tags:
            total_cols += 1  # Tag total
            total_cols += len(tag_ps_structures[tag])  # PS subtotals
        
        # Row 1: Title
        scm(ws, (1, 1), (1, total_cols), group_title, font20b)
        
        # Row 2: Basic columns + Tag headers
        scm(ws, (2, 1), (3, 1), "姓名", font14b)
        scm(ws, (2, 2), (3, 2), "学号", font14b)
        scm(ws, (2, 3), (3, 3), "总分", font14b)
        
        current_col = 4
        for tag in sorted_tags:
            tag_col_count = 1 + len(tag_ps_structures[tag])  # Tag total + PS subtotals
            tag_start_col = current_col
            
            # Tag header (row 2)
            scm(ws, (2, current_col), (2, current_col + tag_col_count - 1), f"{tag}总分", font14b)
            
            # Tag total column (row 3)
            sc(ws, (3, current_col), "小计", font12b)
            current_col += 1
            
            # PS subtotal columns (row 3)
            for psid, ps_name, _ in tag_ps_structures[tag]:
                sc(ws, (3, current_col), ps_name[:15], font10)  # Truncate long names
                current_col += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        # Fill student data with formulas
        for row_idx, (uname, _) in enumerate(member_list, start=4):
            data = student_scores[uname]
            sc(ws, (row_idx, 1), data["nickname"], font10)
            sc(ws, (row_idx, 2), uname, font10)
            
            # Total score - sum of all tag totals
            total_formula_parts = []
            
            current_col = 4
            for tag in sorted_tags:
                tag_start_col = current_col
                tag_terms = []
                
                # Skip tag total column
                current_col += 1
                
                # PS subtotals - reference from detail sheets
                for psid, ps_name, _ in tag_ps_structures[tag]:
                    # Create safe sheet name
                    sheet_name = self._sanitize_sheet_name(ps_name, psid)
                    
                    # Formula: reference subtotal column from detail sheet for this student
                    # Use VLOOKUP to find student by username in column B
                    ps_formula = f"=IFERROR(VLOOKUP(B{row_idx},'{sheet_name}'!B:C,2,FALSE),0)"
                    cell = ws[gc(row_idx, current_col)]
                    cell.value = ps_formula
                    cell.font = font10
                    
                    ps_weight = ps_weight_map.get(psid, 0)
                    try:
                        ps_weight_val = float(ps_weight)
                    except (TypeError, ValueError):
                        ps_weight_val = 0.0
                    tag_terms.append(f"{gc(row_idx, current_col)}/100*{ps_weight_val}")
                    current_col += 1
                
                # Tag total - sum of all PS subtotals in this tag
                if tag_terms:
                    tag_formula = "=" + "+".join(tag_terms)
                    cell = ws[gc(row_idx, tag_start_col)]
                    cell.value = tag_formula
                    cell.font = font10
                    total_formula_parts.append(gc(row_idx, tag_start_col))
                else:
                    sc(ws, (row_idx, tag_start_col), 0, font10)
            
            # Overall total - sum of all tag totals
            if total_formula_parts:
                total_formula = "=" + "+".join(total_formula_parts)
                cell = ws[gc(row_idx, 3)]
                cell.value = total_formula
                cell.font = font12b
            else:
                sc(ws, (row_idx, 3), 0, font12b)
    
    def _create_detail_sheet(self, ws, ps_name, psid, member_list, student_scores,
                           problems, font10, font12b, font14b):
        """Create detailed sheet for a specific problem set (percentage scores)."""
        current_row = 1
        total_cols = 3 + len(problems)
        scm(ws, (current_row, 1), (current_row, total_cols), ps_name, font14b)
        current_row += 1

        # Column headers
        scm(ws, (current_row, 1), (current_row + 1, 1), "姓名", font12b)
        scm(ws, (current_row, 2), (current_row + 1, 2), "学号", font12b)
        scm(ws, (current_row, 3), (current_row + 1, 3), "小计", font12b)

        current_col = 4
        prob_col_map = {}

        for prob in problems:
            # Problem name header
            scm(ws, (current_row, current_col), (current_row, current_col), prob['name'][:20], font12b)

            # Sub-headers
            sc(ws, (current_row + 1, current_col), "题目得分", font10)

            key = prob.get('key', f"{prob.get('gid', 0)}-{prob.get('pid', 0)}")
            prob_col_map[key] = {
                'raw_score': current_col
            }
            current_col += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        for col in range(4, total_cols + 1):
            ws.column_dimensions[number_to_excel_column(col)].width = 12

        # Student data rows
        current_row += 2
        for uname, _ in member_list:
            data = student_scores[uname]
            ps_data = data["scores"].get(psid, {"total": 0, "submissions": {}})

            sc(ws, (current_row, 1), data["nickname"], font10)
            sc(ws, (current_row, 2), uname, font10)

            # Problem columns
            formula_parts = []
            for prob in problems:
                key = prob.get('key', f"{prob.get('gid', 0)}-{prob.get('pid', 0)}")
                cols = prob_col_map.get(key, {})
                if not cols:
                    continue

                sub_data = ps_data.get("submissions", {}).get(key, {})

                raw_score = sub_data.get('score', 0)
                sc(ws, (current_row, cols['raw_score']), raw_score, font10)

                formula_parts.append(gc(current_row, cols['raw_score']))

            # Total
            if formula_parts:
                cell = ws[gc(current_row, 3)]
                cell.value = "=" + "+".join(formula_parts)
                cell.font = font12b
            else:
                sc(ws, (current_row, 3), 0, font12b)

            current_row += 1


__all__ = ["SummaryReportTask"]
